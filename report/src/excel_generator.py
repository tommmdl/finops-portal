"""
Gera o Excel mensal de FinOps com 8 abas.
"""
import io
from datetime import date
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
C_DARK   = "1F4E79"
C_MID    = "2E75B6"
C_LIGHT  = "BDD7EE"
C_ALT    = "D6E4F0"
C_WHITE  = "FFFFFF"
C_ORANGE = "F4B942"

CHART_COLORS = ["#4472C4","#ED7D31","#FFC000","#70AD47","#5A9BD4",
                "#FF0000","#7030A0","#C00000","#00B0F0","#92D050"]

MONEY_FMT = '"$"#,##0.00'
PCT_FMT   = '0.00%'

MESES_PT = {
    "January":"Janeiro","February":"Fevereiro","March":"Março",
    "April":"Abril","May":"Maio","June":"Junho","July":"Julho",
    "August":"Agosto","September":"Setembro","October":"Outubro",
    "November":"Novembro","December":"Dezembro",
}


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _side(s="thin"):
    return Side(style=s)

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _set_header(cell, bg=C_DARK, size=10):
    cell.fill = _fill(bg)
    cell.font = Font(color=C_WHITE, bold=True, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def _set_data(cell, alt=False, bold=False, align="right", fmt=None, color=None):
    if alt:
        cell.fill = _fill(C_ALT)
    cell.font = Font(bold=bold, size=10, color=color or "000000")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if fmt:
        cell.number_format = fmt

def _autowidth(ws, min_w=12, max_w=40):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(length + 3, min_w), max_w)

def _val(row, col):
    """Lê valor de uma linha do iterrows pelo nome da coluna."""
    return row[col] if col in row.index else 0.0

def _fmt_brl(value: float) -> str:
    """Formata em notação brasileira: $29.627,90"""
    s = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"${s}" if value >= 0 else f"-${s}"

def _last_3_months(reference: date) -> list[str]:
    """Retorna os 3 meses terminando no mês de referência (inclusive)."""
    labels = []
    d = date(reference.year, reference.month, 1)
    for _ in range(3):
        month_en, year = d.strftime("%B %Y").split(" ")
        labels.insert(0, f"{MESES_PT[month_en]} {year}")
        d = d - relativedelta(months=1)
    return labels


# ── Gráficos matplotlib ───────────────────────────────────────────────────────

def _buf_from_fig(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    return buf


def _chart_evolucao(totals: dict, month_cols: list[str]) -> io.BytesIO:
    months = month_cols
    values = [totals.get(m, 0) for m in months]
    max_v  = max(values) if values else 1

    fig, ax = plt.subplots(figsize=(11, 5), facecolor="white")
    ax.set_facecolor("white")
    bars = ax.bar(months, values, color="#4472C4", width=0.55, zorder=3)

    ax.spines[["top", "right"]].set_visible(False)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: _fmt_brl(x)))
    ax.tick_params(axis="x", labelsize=9)
    ax.tick_params(axis="y", labelsize=9)
    ax.yaxis.grid(True, linestyle="--", alpha=0.4, zorder=0)
    ax.set_ylim(0, max_v * 1.22)

    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max_v * 0.015,
            _fmt_brl(val),
            ha="center", va="bottom", fontsize=8.5,
            color="#4472C4", fontweight="bold"
        )

    fig.tight_layout()
    return _buf_from_fig(fig)


def _chart_top10(top10: pd.DataFrame, last3: list[str]) -> io.BytesIO:
    """
    Eixo X = meses; dentro de cada mês, uma barra por serviço (cor única por serviço).
    Replica o layout do exemploCHART_ABA4.
    """
    services = top10["Service"].tolist()
    n_svc    = len(services)
    n_months = len(last3)

    # Largura de cada barra e posições dos grupos
    w       = 0.7 / n_svc          # barras cabem em ~0.7 de unidade
    x_base  = list(range(n_months))

    fig, ax = plt.subplots(figsize=(13, 5.5), facecolor="white")
    ax.set_facecolor("white")

    svc_colors = ["#1F4E79","#4472C4","#00B0F0","#375623","#7030A0",
                  "#FFC000","#ED7D31","#FF0000","#E6B8A2","#C00000"]

    for i, (svc, color) in enumerate(zip(services, svc_colors)):
        offsets = [xi + (i - n_svc / 2 + 0.5) * w for xi in x_base]
        vals    = [float(top10.loc[top10["Service"] == svc, m].values[0])
                   if m in top10.columns else 0 for m in last3]
        ax.bar(offsets, vals, width=w * 0.92, label=svc, color=color, zorder=3)

    ax.set_xticks(x_base)
    ax.set_xticklabels(last3, fontsize=10)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"${v:,.2f}"))
    ax.tick_params(axis="y", labelsize=9)
    ax.spines[["top", "right"]].set_visible(False)
    ax.yaxis.grid(True, linestyle="--", alpha=0.4, zorder=0)
    ax.legend(fontsize=7.5, loc="lower center",
              bbox_to_anchor=(0.5, -0.28), ncol=5, frameon=False)

    fig.tight_layout()
    return _buf_from_fig(fig)


# ── Abas ──────────────────────────────────────────────────────────────────────

def _aba1_dados_base(wb: Workbook, df: pd.DataFrame, month_cols: list[str]):
    ws = wb.create_sheet("Dados Base")
    cols = ["Service"] + month_cols

    for c, col in enumerate(cols, 1):
        _set_header(ws.cell(row=1, column=c, value=col))

    for r, (_, row) in enumerate(df.iterrows(), 2):
        alt = r % 2 == 0
        cell = ws.cell(row=r, column=1, value=row["Service"])
        _set_data(cell, alt=alt, align="left")
        for c, m in enumerate(month_cols, 2):
            cell = ws.cell(row=r, column=c, value=float(_val(row, m)))
            _set_data(cell, alt=alt, fmt=MONEY_FMT)

    _autowidth(ws)


def _aba2_detalhamento(wb: Workbook, df: pd.DataFrame, month_cols: list[str]):
    ws = wb.create_sheet("Detalhamento 6 meses")
    cols = ["Service"] + month_cols + ["Média", "Aumento"]
    n    = len(month_cols)

    for c, col in enumerate(cols, 1):
        _set_header(ws.cell(row=1, column=c, value=col))

    for r, (_, row) in enumerate(df.iterrows(), 2):
        alt = r % 2 == 0
        cell = ws.cell(row=r, column=1, value=row["Service"])
        _set_data(cell, alt=alt, align="left")

        for c, m in enumerate(month_cols, 2):
            cell = ws.cell(row=r, column=c, value=float(_val(row, m)))
            _set_data(cell, alt=alt, fmt=MONEY_FMT)

        # Média = AVERAGE dos 6 meses
        b_col  = get_column_letter(2)
        g_col  = get_column_letter(n + 1)
        avg_col = n + 2
        inc_col = n + 3

        avg_cell = ws.cell(row=r, column=avg_col,
                           value=f"=AVERAGE({b_col}{r}:{g_col}{r})")
        _set_data(avg_cell, alt=alt, fmt=MONEY_FMT)

        # Aumento = último mês / média - 1
        last_letter = get_column_letter(n + 1)
        avg_letter  = get_column_letter(avg_col)
        inc_cell = ws.cell(row=r, column=inc_col,
                           value=f"={last_letter}{r}/{avg_letter}{r}-1")
        _set_data(inc_cell, alt=alt, fmt=PCT_FMT)

    _autowidth(ws)


def _aba3_evolucao(wb: Workbook, totals: dict, month_cols: list[str]):
    ws = wb.create_sheet("Evolução Custos AWS - 6 meses")
    cols = ["Service"] + month_cols

    for c, col in enumerate(cols, 1):
        _set_header(ws.cell(row=1, column=c, value=col))

    cell = ws.cell(row=2, column=1, value="Total costs")
    _set_data(cell, bold=True, align="left")
    for c, m in enumerate(month_cols, 2):
        cell = ws.cell(row=2, column=c, value=totals.get(m, 0.0))
        _set_data(cell, bold=True, fmt=MONEY_FMT)

    # Gráfico
    chart_buf = _chart_evolucao(totals, month_cols)
    img = XLImage(chart_buf)
    img.width  = 750
    img.height = 370
    ws.add_image(img, "A4")

    _autowidth(ws)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20


def _aba4_top10(wb: Workbook, df: pd.DataFrame,
                month_cols: list[str], reference: date):
    ws   = wb.create_sheet("Top Ten Serviços")
    last3 = _last_3_months(reference)
    last3 = [m for m in last3 if m in df.columns]

    top10 = df.nlargest(10, last3[-1] if last3 else month_cols[-1])

    cols = ["Service"] + last3
    for c, col in enumerate(cols, 1):
        _set_header(ws.cell(row=1, column=c, value=col))

    for r, (_, row) in enumerate(top10.iterrows(), 2):
        alt = r % 2 == 0
        cell = ws.cell(row=r, column=1, value=row["Service"])
        _set_data(cell, alt=alt, align="left")
        for c, m in enumerate(last3, 2):
            cell = ws.cell(row=r, column=c, value=float(_val(row, m)))
            _set_data(cell, alt=alt, fmt=MONEY_FMT)

    # Gráfico
    chart_buf = _chart_top10(top10.reset_index(drop=True), last3)
    img = XLImage(chart_buf)
    img.width  = 900
    img.height = 450
    ws.add_image(img, "A13")

    _autowidth(ws)


def _aba5_oscilacao(wb: Workbook, df: pd.DataFrame, month_cols: list[str]):
    ws = wb.create_sheet("Serviços com maior oscilação")
    cols = ["Service"] + month_cols + ["Média", "Aumento"]

    # Calcula média e aumento
    df = df.copy()
    df["_avg"] = df[month_cols].mean(axis=1)
    df["_inc"] = df.apply(
        lambda r: (r[month_cols[-1]] / r["_avg"] - 1) if r["_avg"] != 0 else 0,
        axis=1
    )

    aumentos = df[df["_inc"] > 0.30].sort_values("_inc", ascending=False)
    recuos   = df[df["_inc"] < 0].sort_values("_inc", ascending=True)

    def _write_section(start_row: int, data: pd.DataFrame) -> int:
        # Cabeçalho
        for c, col in enumerate(cols, 1):
            _set_header(ws.cell(row=start_row, column=c, value=col))
        r = start_row + 1
        for _, row in data.iterrows():
            cell = ws.cell(row=r, column=1, value=row["Service"])
            _set_data(cell, bold=True, align="left")
            for c, m in enumerate(month_cols, 2):
                cell = ws.cell(row=r, column=c, value=float(_val(row, m)))
                _set_data(cell, fmt=MONEY_FMT)
            # Média
            avg_col = len(month_cols) + 2
            inc_col = len(month_cols) + 3
            cell = ws.cell(row=r, column=avg_col, value=round(float(row["_avg"]), 2))
            _set_data(cell, fmt=MONEY_FMT)
            # Aumento
            cell = ws.cell(row=r, column=inc_col, value=round(float(row["_inc"]), 4))
            _set_data(cell, fmt=PCT_FMT,
                      color="C00000" if row["_inc"] > 0 else "375623")
            r += 1
        return r

    row_cursor = 1
    if not aumentos.empty:
        row_cursor = _write_section(row_cursor, aumentos)
        row_cursor += 2  # espaço entre seções

    if not recuos.empty:
        _write_section(row_cursor, recuos)

    _autowidth(ws)


def _aba6_cobertura(wb: Workbook, reference: date):
    ws = wb.create_sheet("Cobertura Atual de Reservas")

    from dateutil.relativedelta import relativedelta
    months = []
    d = date(reference.year, reference.month, 1)
    for _ in range(3):
        d = d - relativedelta(months=1)
        month_en = d.strftime("%B")
        months.insert(0, MESES_PT[month_en])

    servicos = ["Compute Savings Plans", "RDS"]

    # Tabela Cobertura (col 1-4)
    for c, h in enumerate(["Cobertura"] + months, 1):
        _set_header(ws.cell(row=1, column=c, value=h))
    for r, svc in enumerate(servicos, 2):
        cell = ws.cell(row=r, column=1, value=svc)
        _set_data(cell, align="left")
        for c in range(2, 5):
            _set_data(ws.cell(row=r, column=c, value=""), fmt=PCT_FMT)

    # Tabela Utilização (col 6-9, separada por coluna vazia)
    for c, h in enumerate(["Utilização"] + months, 6):
        _set_header(ws.cell(row=1, column=c, value=h))
    for r, svc in enumerate(servicos, 2):
        cell = ws.cell(row=r, column=6, value=svc)
        _set_data(cell, align="left")
        for c in range(7, 10):
            _set_data(ws.cell(row=r, column=c, value=""), fmt=PCT_FMT)

    # Tabela Economia Acumulada (linha 5+)
    start = 5
    headers_econ = ["Economia Acumulada", "Economia $", "Economia R$"]
    for c, h in enumerate(headers_econ, 1):
        _set_header(ws.cell(row=start, column=c, value=h))

    for r, svc in enumerate(servicos, start + 1):
        cell = ws.cell(row=r, column=1, value=svc)
        _set_data(cell, align="left")
        _set_data(ws.cell(row=r, column=2, value=""), fmt=MONEY_FMT)
        _set_data(ws.cell(row=r, column=3, value=""), fmt='"R$"#,##0.00')

    # Total
    tr = start + len(servicos) + 1
    cell = ws.cell(row=tr, column=1, value="Total")
    _set_data(cell, bold=True, align="left")
    b_col = get_column_letter(2)
    e_col = get_column_letter(2)
    _set_data(ws.cell(row=tr, column=2,
                      value=f"=SUM({b_col}{start+1}:{b_col}{tr-1})"),
              bold=True, fmt=MONEY_FMT)
    _set_data(ws.cell(row=tr, column=3,
                      value=f"=SUM(C{start+1}:C{tr-1})"),
              bold=True, fmt='"R$"#,##0.00')

    _autowidth(ws)


def _aba7_economia(wb: Workbook, usd_brl: float):
    ws = wb.create_sheet("Economia - Aquisição Reservas")
    headers = ["Reserva", "Economia Mês", "Economia Anual", "Economia Anual em R$"]
    servicos = ["Compute Savings Plans", "RDS", "ElastiCache", "OpenSearch"]

    for c, h in enumerate(headers, 1):
        _set_header(ws.cell(row=1, column=c, value=h))

    for r, svc in enumerate(servicos, 2):
        cell = ws.cell(row=r, column=1, value=svc)
        _set_data(cell, align="left")
        _set_data(ws.cell(row=r, column=2, value=0.0), fmt=MONEY_FMT)
        cell_c = ws.cell(row=r, column=3, value=f"=B{r}*12")
        _set_data(cell_c, fmt=MONEY_FMT)
        cell_d = ws.cell(row=r, column=4, value=f"=C{r}*{usd_brl}")
        _set_data(cell_d, fmt='"R$"#,##0.00')

    # Total
    tr = len(servicos) + 2
    cell = ws.cell(row=tr, column=1, value="Total")
    _set_data(cell, bold=True, align="left")
    for c, fmt in [(2, MONEY_FMT), (3, MONEY_FMT), (4, '"R$"#,##0.00')]:
        col = get_column_letter(c)
        cell = ws.cell(row=tr, column=c,
                       value=f"=SUM({col}2:{col}{tr-1})")
        _set_data(cell, bold=True, fmt=fmt)

    _autowidth(ws)


def _aba8_cloudcheckr(wb: Workbook):
    ws = wb.create_sheet("Pontos de Atenção - CloudCheckr")

    # Seção texto
    ws.cell(row=1, column=1, value="Recurso - Economia").font = \
        Font(bold=True, size=11, color=C_DARK)
    ws.merge_cells("A1:C1")
    for r in range(2, 10):
        cell = ws.cell(row=r, column=1, value="")
        cell.font = Font(size=10)

    ws.cell(row=10, column=1, value="Economia Total:").font = \
        Font(bold=True, size=10)
    ws.cell(row=10, column=2, value=0.0).number_format = MONEY_FMT

    # Espaço
    # Seção tabela
    start = 12
    for c, h in enumerate(["Recurso", "Quantidade", "Economia ($)"], 1):
        _set_header(ws.cell(row=start, column=c, value=h))

    for r in range(start + 1, start + 9):
        _set_data(ws.cell(row=r, column=1, value=""), align="left")
        _set_data(ws.cell(row=r, column=2, value=""), align="center")
        _set_data(ws.cell(row=r, column=3, value=""), fmt=MONEY_FMT)

    _autowidth(ws)


# ── Entry point ───────────────────────────────────────────────────────────────

def generate(
    costs_df: pd.DataFrame,
    totals_by_month: dict,
    month_cols: list[str],
    client_name: str,
    reference: date,
    usd_brl: float,
    output_dir: Path,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _aba1_dados_base(wb, costs_df, month_cols)
    _aba2_detalhamento(wb, costs_df, month_cols)
    _aba3_evolucao(wb, totals_by_month, month_cols)
    _aba4_top10(wb, costs_df, month_cols, reference)
    _aba5_oscilacao(wb, costs_df, month_cols)
    _aba6_cobertura(wb, reference)
    _aba7_economia(wb, usd_brl)
    _aba8_cloudcheckr(wb)

    month_en, year = reference.strftime("%B %Y").split(" ")
    from extractor import MESES_PT
    month_pt = f"{MESES_PT[month_en]} {year}"
    filename = f"{client_name} _ Report Mensal {month_pt}.xlsx"
    path = output_dir / filename
    wb.save(path)
    return path
